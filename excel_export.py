"""
Excel Export Module
Генерирует красиво отформатированный Excel файл с результатами парсинга
ОБНОВЛЕНО: Одна колонка Material (всегда из PDF)
"""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def auto_adjust_column_width(ws, min_width: int = 10, max_width: int = 50):
    """
    Автоматически подстраивает ширину колонок под содержимое

    Args:
        ws: worksheet
        min_width: минимальная ширина колонки
        max_width: максимальная ширина колонки
    """

    for column in ws.columns:
        max_length = 0
        column_letter = None

        for cell in column:
            # Пропускаем merged cells
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue

            column_letter = cell.column_letter

            try:
                # Вычисляем длину текста в ячейке
                if cell.value:
                    cell_value = str(cell.value)
                    # Учитываем переносы строк
                    lines = cell_value.split("\n")
                    current_max = max(len(line) for line in lines)

                    if current_max > max_length:
                        max_length = current_max
            except:
                pass

        # Устанавливаем ширину с учетом границ
        if column_letter:
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width


def generate_excel_report(data: dict, output_path: str = None) -> str:
    """
    Генерирует Excel отчёт с цветовой индикацией

    НОВАЯ СТРУКТУРА:
    - Только одна колонка Material (из PDF - истина!)
    - Status: equal / notEqual / new
    - Цветовая индикация по статусу

    Args:
        data: результат парсинга (table1, table2, table3)
        output_path: путь для сохранения (если None - создаёт временный)

    Returns:
        str: путь к созданному файлу
    """

    # Создаём новую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parsing Results"

    # Цвета для статусов
    COLOR_EQUAL = "C6EFCE"  # Светло-зелёный
    COLOR_NOT_EQUAL = "FFC7CE"  # Светло-красный
    COLOR_NEW_ITEM = "FFEB9C"  # Светло-жёлтый
    COLOR_HEADER = "4472C4"  # Синий

    # Стили
    header_fill = PatternFill(
        start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)

    equal_fill = PatternFill(
        start_color=COLOR_EQUAL, end_color=COLOR_EQUAL, fill_type="solid"
    )
    not_equal_fill = PatternFill(
        start_color=COLOR_NOT_EQUAL, end_color=COLOR_NOT_EQUAL, fill_type="solid"
    )
    new_item_fill = PatternFill(
        start_color=COLOR_NEW_ITEM, end_color=COLOR_NEW_ITEM, fill_type="solid"
    )

    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    # ========== ЗАГОЛОВОК ==========
    ws.merge_cells("A1:G1")
    ws["A1"] = "PDF PARSING RESULTS"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = center_alignment

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(size=10, italic=True)
    ws["A2"].alignment = center_alignment

    # ========== СТАТИСТИКА ==========
    table2 = data.get("table2", [])

    total = len(table2)
    equal = len([c for c in table2 if c.get("status") == "equal"])
    not_equal = len([c for c in table2 if c.get("status") == "notEqual"])
    new_items = len([c for c in table2 if c.get("status") == "new"])

    ws["A4"] = "STATISTICS:"
    ws["A4"].font = Font(bold=True, size=12)

    stats_data = [
        ("Total Components:", total),
        ("✅ Equal:", equal),
        ("❌ Not Equal:", not_equal),
        ("🆕 New Items:", new_items),
    ]

    for idx, (label, value) in enumerate(stats_data, start=5):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = value
        ws[f"A{idx}"].font = Font(bold=True)

    # ========== ТАБЛИЦА ЗАГОЛОВКИ ==========
    # НОВОЕ: Только одна колонка Material (из PDF)!
    headers = [
        "Pos",
        "Description",
        "Material",  # ← ОДНА колонка! (из PDF - истина)
        "Quantity",
        "Manager Quantity",  # ← Добавляем если есть
        "Status",
        "Note",
    ]
    header_row = 10

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    # ========== ДАННЫЕ ==========
    current_row = header_row + 1

    for component in table2:
        pos = component.get("pos", "-")
        description = component.get("description", "")

        # НОВОЕ: material теперь строка (не объект!)
        material = component.get("material", "-")
        if not material or material == "":
            material = "-"

        quantity = component.get("quantity", "-")
        if quantity is None or quantity == "":
            quantity = "-"

        manager_quantity = component.get("manager_quantity", "-")
        if manager_quantity is None or manager_quantity == "":
            manager_quantity = "-"

        status = component.get("status", "-")
        note = component.get("note", "-")

        # Определяем цветовой фон по статусу
        if status == "new":
            status_text = "🆕 New Item"
            row_fill = new_item_fill
        elif status == "notEqual":
            status_text = "❌ Not Equal"
            row_fill = not_equal_fill
        elif status == "equal":
            status_text = "✅ Equal"
            row_fill = equal_fill
        else:
            status_text = status
            row_fill = None

        # Заполняем строку
        row_data = [
            pos,
            description,
            material,  # ← Всегда из PDF (истина!)
            quantity,
            manager_quantity,
            status_text,
            note,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.border = border

            if col_idx == 1:  # Pos
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

            # Применяем цветовой фон
            if row_fill:
                cell.fill = row_fill

        current_row += 1

    # ========== ЛЕГЕНДА ==========
    legend_row = current_row + 2

    ws[f"A{legend_row}"] = "LEGEND:"
    ws[f"A{legend_row}"].font = Font(bold=True, size=11)

    legend_data = [
        ("✅ Equal", "Materials match (smart comparison)", equal_fill),
        ("❌ Not Equal", "Materials do not match", not_equal_fill),
        ("🆕 New Item", "Found only in Manager Excel", new_item_fill),
    ]

    for idx, (icon, description, fill) in enumerate(legend_data, start=legend_row + 1):
        ws[f"A{idx}"] = icon
        ws[f"B{idx}"] = description
        ws[f"A{idx}"].fill = fill
        ws[f"A{idx}"].border = border
        ws[f"B{idx}"].border = border

    # ========== ШИРИНА КОЛОНОК (автоматическая) ==========
    auto_adjust_column_width(ws, min_width=8, max_width=50)

    # ========== СОХРАНЕНИЕ ==========
    if output_path is None:
        output_path = (
            f"/tmp/parsing_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

    wb.save(output_path)

    return output_path


def generate_excel_from_api_response(
    api_response: dict, output_path: str = None
) -> str:
    """
    Wrapper для генерации Excel из API response

    Args:
        api_response: полный ответ от API с success, data, validation
        output_path: путь для сохранения

    Returns:
        str: путь к созданному файлу
    """

    if not api_response.get("success"):
        raise ValueError("Cannot generate Excel from failed API response")

    data = api_response.get("data", {})

    return generate_excel_report(data, output_path)
