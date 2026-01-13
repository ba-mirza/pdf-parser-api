import openpyxl
from fuzzywuzzy import fuzz

from hybrid_compare import smart_material_match


def parse_bom_sheet(filepath, sheet_index):
    """
    Парсит конкретный лист BOM.xlsx

    ИСПРАВЛЕНИЯ:
    1. Обрабатывает пустые строки с полным материалом (они относятся к строке выше)
    2. Использует полный материал если доступен

    Returns:
        {
            "size": "12\"",
            "asme": "ASME 600",
            "ends": "RTJ",
            "components": {
                "Body": {
                    "quantity": 1,
                    "material": "ASTM A350 LF2 CL.1"  ← Полный материал!
                },
                ...
            }
        }
    """

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[sheet_index]

    # Извлекаем поля для валидации из строки 28
    size = ws.cell(28, 4).value  # D28 - Dimensione
    asme = ws.cell(28, 15).value  # O28 - Classe
    ends = ws.cell(28, 25).value  # Y28 - Estremità

    components = {}

    # Переменная для хранения последнего компонента
    last_component_name = None
    last_component_data = None

    # Парсим компоненты (строки 31+)
    for row in ws.iter_rows(min_row=31, max_row=100, values_only=True):
        if len(row) <= 42:
            continue

        component_name = row[28]  # AC (index 28) - Descrizione componente
        quantity = row[42]  # AQ (index 42) - Q.tà
        material = row[5]  # F (index 5) - MAT.

        # СЛУЧАЙ 1: Строка с компонентом
        if component_name and str(component_name).strip() != "":
            name = str(component_name).strip()

            # Пропускаем заголовок
            if name == "Descrizione componente":
                continue

            # Сохраняем последний компонент (на случай если следующая строка пустая)
            last_component_name = name

            try:
                qty = int(quantity) if quantity else 1
            except (ValueError, TypeError):
                qty = 1

            # Извлекаем материал из колонки F (MAT.)
            mat = str(material).strip() if material else ""

            last_component_data = {"quantity": qty, "material": mat}

            # Сохраняем компонент
            components[name] = last_component_data

        # СЛУЧАЙ 2: Пустая строка (component_name пустое, но есть материал)
        elif material and str(material).strip() != "" and last_component_name:
            # Эта строка относится к предыдущему компоненту!
            full_material = str(material).strip()

            # Обновляем материал предыдущего компонента ПОЛНЫМ названием
            if last_component_name in components:
                components[last_component_name]["material"] = full_material
                print(f"  ✅ Обновлен материал для '{last_component_name}': '{full_material}'")

    return {
        "size": str(size).strip() if size else None,
        "asme": str(asme).strip() if asme else None,
        "ends": str(ends).strip() if ends else None,
        "components": components,
    }


def validate_bom_with_pdf(bom_data, pdf_size, pdf_asme, pdf_ends):
    """
    Валидирует что BOM соответствует PDF

    Returns:
        {
            "valid": True/False,
            "errors": [...]
        }
    """

    errors = []

    # Нормализуем для сравнения
    bom_size = (
        bom_data["size"].replace('"', "").replace("'", "").strip()
        if bom_data["size"]
        else ""
    )
    pdf_size_clean = (
        pdf_size.replace('"', "").replace("'", "").strip() if pdf_size else ""
    )

    bom_asme = bom_data["asme"].strip() if bom_data["asme"] else ""
    pdf_asme_clean = pdf_asme.strip() if pdf_asme else ""

    bom_ends = bom_data["ends"].strip() if bom_data["ends"] else ""
    pdf_ends_clean = pdf_ends.strip() if pdf_ends else ""

    # Проверяем совпадение
    if bom_size not in pdf_size_clean and pdf_size_clean not in bom_size:
        errors.append(f"SIZE не совпадает: BOM={bom_data['size']}, PDF={pdf_size}")

    if bom_asme not in pdf_asme_clean and pdf_asme_clean not in bom_asme:
        errors.append(f"ASME не совпадает: BOM={bom_data['asme']}, PDF={pdf_asme}")

    if bom_ends not in pdf_ends_clean and pdf_ends_clean not in bom_ends:
        errors.append(f"ENDS не совпадает: BOM={bom_data['ends']}, PDF={pdf_ends}")

    return {"valid": len(errors) == 0, "errors": errors}


def parse_manager_sheet(filepath, target_size, target_asme):
    """
    Парсит order-manager.xlsx (только первый лист)
    Ищет строку где Size = target_size И Class = target_asme

    Returns:
        {
            "found": True,
            "row": 15,
            "materials": {
                "Body": "A350 LF2",
                "Closures": "A350 LF2",
                "Ball": "A182 F316/F316L",
                ...
            }
        }
    """

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active  # Первый лист

    # Маппинг колонок на компоненты (Col 18-30)
    component_columns = {
        18: "Body",
        19: "Closures",
        20: "Gland",
        21: "Trunnion",
        22: "Weld overlay",
        23: "Ball",
        24: "Seat rings",
        25: "Seat inserts",
        26: "Stem",
        27: "Dynamic seals",
        28: "Static seals",
        29: "Fire Safe gaskets",
        30: "Springs",
    }

    # Нормализуем target для сравнения
    target_size_clean = (
        target_size.replace('"', "").replace("'", "").strip() if target_size else ""
    )
    target_asme_clean = target_asme.strip() if target_asme else ""

    # Ищем строку с нужными Size и Class
    for row_idx in range(15, 100):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

        if len(row) < 30:
            continue

        size_value = row[8]  # Col 9 - Size (Inch)
        class_value = row[9]  # Col 10 - Class (lbs)

        if not size_value or not class_value:
            continue

        # Нормализуем значения
        size_str = str(size_value).replace('"', "").replace("'", "").strip()
        class_str = str(class_value).strip()

        # Проверяем совпадение
        if target_size_clean in size_str and target_asme_clean in class_str:
            # Нашли нужную строку! Извлекаем материалы
            materials = {}

            for col_idx, component_name in component_columns.items():
                material = row[col_idx - 1]  # -1 потому что row с индекса 0

                if material and str(material).strip():
                    materials[component_name] = str(material).strip()

            return {
                "found": True,
                "row": row_idx,
                "size": size_str,
                "asme": class_str,
                "materials": materials,
            }

    return {
        "found": False,
        "error": f"Не найдена строка с Size={target_size} и Class={target_asme}",
    }


def find_matching_manager_column(pdf_description, manager_materials):
    """
    Находит соответствующую колонку Manager для компонента PDF

    ИСПРАВЛЕНО: Возвращает ЛУЧШЕЕ совпадение, а не первое!

    Args:
        pdf_description: "Body", "Ball", "Seat Spring", etc.
        manager_materials: {"Body": "A350 LF2", "Ball": "A182 F316", ...}

    Returns:
        ("Body", "A350 LF2") или (None, None)
    """

    pdf_lower = pdf_description.lower().strip()

    # 1. Точное совпадение (case-insensitive)
    for manager_key, manager_value in manager_materials.items():
        if pdf_lower == manager_key.lower():
            return (manager_key, manager_value)

    # 2. Fuzzy match - НО БЕРЕМ ЛУЧШЕЕ, А НЕ ПЕРВОЕ!
    best_match = None
    best_score = 0

    for manager_key, manager_value in manager_materials.items():
        score = fuzz.token_sort_ratio(pdf_lower, manager_key.lower())

        if score > best_score and score >= 80:  # Порог 80
            best_score = score
            best_match = (manager_key, manager_value)

    return best_match if best_match else (None, None)


def merge_all_data(pdf_data, bom_components, manager_materials):
    """
    Объединяет данные из PDF, BOM и Manager с УМНЫМ сравнением материалов

    ИСПРАВЛЕНИЯ:
    1. Используем ЛУЧШЕЕ совпадение fuzzy search, а не первое
    2. BOM теперь содержит полные материалы благодаря parse_bom_sheet

    Структура ответа:
    {
        "pos": "10",
        "description": "Body",
        "material": "ASTM A350 LF2 CL1",        ← Всегда из PDF!
        "bom_material": "A350 LF2",             ← Для UI
        "order_material": "A350 LF2",           ← Для UI
        "quantity": 2,                          ← Из BOM
        "manager_quantity": None,
        "status": "equal",                      ← equal / notEqual / new
        "note": ""
    }
    """

    # Отслеживаем какие колонки Manager мы уже использовали
    used_manager_columns = set()

    # Обрабатываем компоненты из PDF
    for item in pdf_data.get("table2", []):
        description = item.get("description", "").strip()
        pdf_material = item.get("material", "").strip()

        if not description:
            continue

        # ===== ШАГ 1: Ищем в BOM (ИСПРАВЛЕНО!) =====
        matched_bom = None
        bom_material = ""

        # Ищем ЛУЧШЕЕ совпадение, а не первое!
        best_match_name = None
        best_match_score = 0
        best_match_data = None

        for bom_name, bom_data in bom_components.items():
            score = fuzz.token_sort_ratio(description.lower(), bom_name.lower())

            if score > best_match_score and score >= 70:
                best_match_score = score
                best_match_name = bom_name
                best_match_data = bom_data

        if best_match_data:
            matched_bom = best_match_data
            bom_material = best_match_data.get("material", "")
            print(f"  🔗 '{description}' → '{best_match_name}' (score: {best_match_score})")

        # Достаем quantity из BOM
        bom_quantity = matched_bom["quantity"] if matched_bom else None

        # ===== ШАГ 2: Умное сравнение материала PDF с BOM =====
        bom_materials_match = False
        if pdf_material and bom_material:
            bom_materials_match, _ = smart_material_match(pdf_material, bom_material)

        # ===== ШАГ 3: Находим соответствующую колонку в Manager =====
        manager_column, manager_material = find_matching_manager_column(
            description, manager_materials
        )

        if manager_column:
            used_manager_columns.add(manager_column)

        # ===== ШАГ 4: Умное сравнение материала PDF с Order =====
        order_materials_match = False
        if pdf_material and manager_material:
            order_materials_match, _ = smart_material_match(pdf_material, manager_material)

        # ===== ШАГ 5: Определяем статус =====
        if bom_materials_match or order_materials_match:
            status = "equal"
        elif (bom_material or manager_material) and not (
            bom_materials_match or order_materials_match
        ):
            status = "notEqual"
        else:
            status = "equal"

        # ===== ШАГ 6: Формируем результат =====
        item["material"] = pdf_material
        item["bom_material"] = bom_material if bom_material else None
        item["order_material"] = manager_material if manager_material else None
        item["quantity"] = bom_quantity
        item["manager_quantity"] = None
        item["status"] = status

        if "note" not in item:
            item["note"] = ""

    # ===== ШАГ 7: Добавляем компоненты которых НЕТ в PDF, но ЕСТЬ в Manager =====
    for manager_column, manager_material in manager_materials.items():
        if manager_column not in used_manager_columns:
            new_item = {
                "pos": "",
                "description": manager_column,
                "material": manager_material,
                "bom_material": None,
                "order_material": manager_material,
                "quantity": None,
                "manager_quantity": None,
                "status": "new",
                "note": "",
            }

            pdf_data.get("table2", []).append(new_item)

    return pdf_data
